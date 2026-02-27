from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
import re

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException


@dataclass(frozen=True)
class _Cell:
    value: object


class TabularWorksheet:
    def __init__(self, rows: list[list[object]]) -> None:
        self._rows = rows
        self.max_row = len(rows)
        self.max_column = max((len(row) for row in rows), default=0)

    def cell(self, row: int, column: int) -> _Cell:
        row_idx = row - 1
        col_idx = column - 1
        if row_idx < 0 or col_idx < 0:
            return _Cell(None)
        if row_idx >= len(self._rows):
            return _Cell(None)
        row_values = self._rows[row_idx]
        if col_idx >= len(row_values):
            return _Cell(None)
        return _Cell(row_values[col_idx])


def _dataframe_to_rows(dataframe: pd.DataFrame) -> list[list[object]]:
    normalized = dataframe.where(pd.notna(dataframe), None)
    rows = normalized.values.tolist()
    return [list(row) for row in rows]


def load_excel_sheets(path: str | Path) -> dict[str, object]:
    excel_path = Path(path)
    suffix = excel_path.suffix.lower()

    if suffix == ".csv":
        dataframe = pd.read_csv(excel_path, header=None, dtype=object, encoding_errors="ignore")
        return {"CSV": TabularWorksheet(_dataframe_to_rows(dataframe))}

    if suffix == ".xls":
        return _load_sheets_with_pandas(excel_path)

    try:
        workbook = load_workbook(excel_path, data_only=True)
        return {sheet_name: workbook[sheet_name] for sheet_name in workbook.sheetnames}
    except InvalidFileException:
        return _load_sheets_with_pandas(excel_path)


def _load_sheets_with_pandas(path: Path) -> dict[str, object]:
    excel_file = pd.ExcelFile(path)
    sheets: dict[str, object] = {}
    for sheet_name in excel_file.sheet_names:
        dataframe = excel_file.parse(sheet_name, header=None, dtype=object)
        sheets[sheet_name] = TabularWorksheet(_dataframe_to_rows(dataframe))
    return sheets


def normalize_loose_key(value: str) -> str:
    raw = str(value or "").upper()
    normalized = re.sub(r"[^A-Z0-9]+", " ", raw)
    return " ".join(normalized.split())
